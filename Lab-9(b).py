from openpyxl import Workbook
from openpyxl import Font

wb=Workbook()
sheet=wb.active
sheet.title="Language"
wb.create_sheet(title="Capital")
lang=["Kannada","telugu","Tamil"]
state=["Karnataka","Telangana","Tamil Nadu"]
capital=["bengaluru","hyderbad","chennai"]
code=['KA','TS','TN']
sheet.cell(row=1,column=1).value="State"
sheet.cell(row=1,column=2).value="Language"
sheet.cell(row=1,column=3).value="Code"
ft=Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font=ft
    for i in range(2,5):
        sheet.cell(row=i,column=1).value=state[i-2]
        sheet.cell(row=i,column=2).value=lang[i-2]
        sheet.cell(row=i,column=3).value=code[i-2]
wb.save("hello.xlsx")
sheet=wb["Capital"]
sheet.cell(row=1,column=1).value="State"
sheet.cell(row=1,column=2).value="Capital"
sheet.cell(row=1,column=3).value="Code"
ft=Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font=ft
    for i in range(2,5):
        sheet.cell(row=i,column=1).value=state[i-2]
        sheet.cell(row=i,column=2).value=capital[i-2]
        sheet.cell(row=i,column=3).value=code[i-2]
wb.save("hello.xlsx")
srchCode=input("Enter the state code for finding the capital:")
for i in range(2,5):
    data=sheet.cell(row=i,column=3).value
    if srchCode==data:
        print("Ans:",sheet.cell(row=i,column=2).value)
sheet=wb["Language"]
srchCode=input("Enter the state code for finding the language:")
for i in range(2,5):
    data=sheet.cell(row=i,column=3).value
    if srchCode==data:
        print("Ans:",sheet.cell(row=i,column=2).value)
        


